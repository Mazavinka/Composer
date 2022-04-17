from openpyxl import load_workbook
import os
import os.path
import glob
from notfoundpathwindow import NotFoundPathWindow
from config import config as config

class Parser:

    def __init__(self, path):
        self.path = path
        self.book = load_workbook(self.path, data_only=True)
        self.sheet = self.book.active
        self.server_folder = config.get_server_folder()

    def parse_book(self):
        data = {"result": []}

        for row in range(10, self.sheet.max_row):
            if self.sheet[row][1].value is not None:
                data["result"].append({
                    "img_row": row,
                    "img_number": self.sheet[row][1].value.replace(' ', ''),
                    "img_code": self.sheet[row][2].value.replace('p', '').replace('r', '').replace(' ', ''),
                    "img_palette": str(self.sheet[row][3].value).replace(' ', ''),
                    "img_type": self.sheet[row][4].value.replace(' ', ''),
                    "img_width": str(self.sheet[row][5].value).replace(' ', '').replace(",", "."),
                    "img_length_roll": str(self.sheet[row][6].value).replace(' ', ''),
                    "img_count_roll": str(self.sheet[row][7].value).replace(' ', ''),
                    "img_full_path": 'Not Found',
                })
            else:
                break

        return data

    def set_base_path_to_image(self):
        data = self.parse_book()
        for i in data["result"]:
            for dir in os.listdir(self.get_palette_folder()):
                if i["img_number"][1:] in dir.split():
                    i['img_full_path'] = os.path.join(self.get_palette_folder(), dir)
                    for j in os.listdir(i["img_full_path"]):
                        if i["img_code"] in j.split():
                            i["img_full_path"] = os.path.join(self.get_palette_folder(), dir, j)
        return data

    def get_full_path_to_image(self):
        result = self.set_base_path_to_image()
        for row in result["result"]:
            for file in glob.glob(row['img_full_path'] + '/**/*.pcx', recursive=True):
                if row['img_type'].lower() == 'покрытие':
                    img_file_url = self.name_template_for_canvas(row)
                    if img_file_url in self.rus_to_eng_letters(file):
                        row['img_full_path'] = file
                elif row['img_type'].lower() == 'дорожка':
                    img_file_url = self.name_template_for_track(row)
                    if img_file_url in self.rus_to_eng_letters(file):
                        row['img_full_path'] = file
                else:
                    continue
        return result


    def rus_to_eng_letters(self, string):
        rus_letters = ['а', 'х', 'р', 'с', 'о', 'е', 'к', 'у']
        eng_letters = ['a', 'x', 'p', 'c', 'o', 'e', 'k', 'y']
        for rus, eng in zip(rus_letters, eng_letters):
            string = string.replace(rus, eng)
        return string.lower()

    def get_palette_number(self):
        palette_number = str(self.sheet['D10'].value).replace(' ', '')
        """config.set_palette_number(palette_number)"""
        return palette_number

    def get_palette_folder(self):
        dir_list = os.listdir(self.server_folder)
        for dir in dir_list:
            if dir == self.get_palette_number():
                palette_folder = os.path.join(self.server_folder, self.get_palette_number())
                return palette_folder
            elif self.get_palette_number() in dir:
                palette_folder = os.path.join(self.server_folder, dir)
                return palette_folder
            else:
                continue

    def name_template_for_canvas(self, row):
        template_name = row['img_number'] + row['img_code'] + "xpx" + self.get_palette_number() + ".pcx"
        return template_name.lower()

    def name_template_for_track(self, row):
        template_name = row['img_number'] + row['img_code'] + "xrx" + self.get_palette_number() + 'x' + \
            str(int(float(row['img_width']) * 100)) + ".pcx"
        return template_name.lower()

    def check_empty_path(self):
        result = self.get_full_path_to_image()
        win_error = NotFoundPathWindow()
        win_error.hide_window()
        for i in result["result"]:
            if i['img_full_path'][-4:].lower() != ".pcx":
                if i['img_type'].lower() == 'дорожка':
                    win_error.show_window()
                    label = win_error.show_label_for_track(i)
                    label.pack()
                    i['img_full_path'] = win_error.open_image(self.get_palette_folder())
                    for j in result["result"]:
                        if i['img_full_path'][-4:].lower() == ".pcx":
                            if i["img_number"] == j["img_number"] and i["img_code"] == j["img_code"] and i[
                                "img_type"] == j["img_type"] and i["img_width"] == j["img_width"]:
                                j["img_full_path"] = i["img_full_path"]
                    label.destroy()
                    continue
                elif i['img_type'].lower() == 'покрытие':
                    win_error.show_window()
                    label = win_error.show_label_for_canvas(i)
                    label.pack()
                    i['img_full_path'] = win_error.open_image(self.get_palette_folder())
                    for j in result["result"]:
                        if i['img_full_path'][-4:].lower() == ".pcx":
                            if i["img_number"] == j["img_number"] and i["img_code"] == j["img_code"] and i[
                                "img_type"] == j["img_type"]:
                                j["img_full_path"] = i["img_full_path"]
                    label.destroy()
                    continue
        win_error.destroy()
        return result




if __name__ == "__main__":
    a = Parser("LM52_12-16.05_117.xlsx")
    """for i in a.get_full_path_to_image()["result"]:
        print(i)

    print(a.get_palette_number())
    print(a.get_palette_folder())"""
    print(a.check_empty_path())
